"""Build the reproducible Urea OTS stream-lag evidence workbook.

The supplied 30-second trend rows are synthetic linear interpolation.  This script retains only
the original hourly anchors, builds hourly gradients with spreadsheet formulas, and writes
integer-hour correlation checks.  It does not infer subhour delay from interpolated samples.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TREND_DIR = Path(
    r"D:\Work\Urea Simulation Docs\New folder\Trends\Trends To Excel"
)
DEFAULT_OUTPUT = ROOT / "docs" / "analysis" / "urea_stream_lag_analysis.xlsx"
TREND_FILES = {
    "Normal": "Urea_NormalOp_29-06-2025_Trends.xlsx",
    "Startup": "Urea_Startup_28-06-2025_Trends.xlsx",
}

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "E2F0D9"
RED = "FCE4D6"
GRAY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="A6A6A6")


def hourly_anchors(path: Path) -> tuple[str, list[str], list[tuple]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    notice = str(rows[0][0] or "")
    headers = [str(value) for value in rows[2] if value is not None]
    data = [tuple(row[: len(headers)]) for row in rows[3:] if row[0] is not None]
    first = data[0][0]
    anchors = [
        row for row in data
        if row[0].minute == first.minute and row[0].second == first.second
    ]
    return notice, headers, anchors


def title(ws, text: str, last_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    cell = ws.cell(1, 1, text)
    cell.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 23


def style_table(ws, header_row: int, min_row: int, max_row: int, max_col: int) -> None:
    for cell in ws[header_row]:
        if cell.column <= max_col:
            cell.font = Font(name="Arial", bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.border = Border(bottom=THIN)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def fit_columns(ws, minimum: int = 10, maximum: int = 42) -> None:
    for column in range(1, ws.max_column + 1):
        width = max(
            (len(str(ws.cell(row, column).value or "")) for row in range(1, ws.max_row + 1)),
            default=minimum,
        )
        ws.column_dimensions[get_column_letter(column)].width = min(max(width + 2, minimum), maximum)


def add_anchor_sheet(wb: Workbook, label: str, headers: list[str], anchors: list[tuple]) -> None:
    ws = wb.create_sheet(f"{label} Anchors")
    title(ws, f"{label} — independent hourly anchors", len(headers))
    ws.cell(2, 1, "Only these rows are independent measurements; 30-second rows were interpolated.")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(2, 1).fill = PatternFill("solid", fgColor=BLUE)
    ws.append([])
    ws.append(headers)
    for row in anchors:
        ws.append(row)
    style_table(ws, 4, 4, ws.max_row, len(headers))
    ws.freeze_panes = "B5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{ws.max_row}"
    for row in range(5, ws.max_row + 1):
        ws.cell(row, 1).number_format = "yyyy-mm-dd hh:mm"
    fit_columns(ws, maximum=24)


def add_gradient_sheet(wb: Workbook, label: str, headers: list[str], anchor_count: int) -> None:
    ws = wb.create_sheet(f"{label} Gradients")
    title(ws, f"{label} — hourly gradients", len(headers))
    ws.cell(2, 1, "Formula: value at hour h minus value at hour h-1.")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.append([])
    ws.append(headers)
    anchor_sheet = f"'{label} Anchors'"
    for interval in range(anchor_count - 1):
        target_row = 5 + interval
        later_anchor = 6 + interval
        earlier_anchor = 5 + interval
        ws.cell(target_row, 1, f"={anchor_sheet}!A{later_anchor}")
        for column in range(2, len(headers) + 1):
            letter = get_column_letter(column)
            ws.cell(
                target_row,
                column,
                f"={anchor_sheet}!{letter}{later_anchor}-{anchor_sheet}!{letter}{earlier_anchor}",
            )
    style_table(ws, 4, 4, ws.max_row, len(headers))
    ws.freeze_panes = "B5"
    for row in range(5, ws.max_row + 1):
        ws.cell(row, 1).number_format = "yyyy-mm-dd hh:mm"
        for column in range(2, len(headers) + 1):
            ws.cell(row, column).number_format = "0.0000"
    fit_columns(ws, maximum=24)


def corr_formula(sheet: str, driver_col: int, response_col: int,
                 first: int, last: int, lag: int) -> str:
    driver_letter = get_column_letter(driver_col)
    response_letter = get_column_letter(response_col)
    driver_last = last - lag
    response_first = first + lag
    return (
        f'=IFERROR(CORREL(\'{sheet}\'!{driver_letter}{first}:{driver_letter}{driver_last},'
        f"'{sheet}'!{response_letter}{response_first}:{response_letter}{last}),\"\")"
    )


def add_lag_estimates(wb: Workbook, datasets: dict[str, dict]) -> None:
    ws = wb.create_sheet("Lag Estimates")
    headers = [
        "Dataset", "Response tag", "Independent anchors", "Independent gradients",
        "Lag 0 h r", "Lag 1 h r", "Lag 2 h r", "Lag 3 h r", "Interpretation",
    ]
    title(ws, "Integer-hour gradient correlation checks", len(headers))
    ws.cell(
        2, 1,
        "These checks can distinguish hourly bins only. Startup has six gradients; longer-lag fits are overfit.",
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.append([])
    ws.append(headers)
    row = 5
    for label, data in datasets.items():
        headers_in = data["headers"]
        anchor_count = len(data["anchors"])
        gradient_count = anchor_count - 1
        driver_col = headers_in.index("UREA-LOAD") + 1
        for response_col, response in enumerate(headers_in[1:], start=2):
            ws.cell(row, 1, label)
            ws.cell(row, 2, response)
            ws.cell(row, 3, anchor_count)
            ws.cell(row, 4, gradient_count)
            for lag in range(4):
                ws.cell(
                    row, 5 + lag,
                    corr_formula(
                        f"{label} Gradients", driver_col, response_col,
                        5, 4 + gradient_count, lag,
                    ),
                )
            ws.cell(
                row, 9,
                f'=IF(OR(D{row}<7,COUNT(E{row}:H{row})<4),"INSUFFICIENT",'
                f'IF(ABS(E{row})>=MAX(ABS(F{row}),ABS(G{row}),ABS(H{row})),"0 h bin","ambiguous"))',
            )
            row += 1
    style_table(ws, 4, 4, ws.max_row, len(headers))
    for data_row in range(5, ws.max_row + 1):
        for column in range(5, 9):
            ws.cell(data_row, column).number_format = "0.0000"
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:I{ws.max_row}"
    fit_columns(ws)


def add_route_parameters(wb: Workbook) -> None:
    sys.path.insert(0, str(ROOT / "backend"))
    import main  # pylint: disable=import-outside-toplevel

    ws = wb.create_sheet("Route Parameters")
    headers = [
        "Route", "Source", "Destination", "Design flow kg/h", "Design dead time s",
        "Effective inventory kg", "Maximum dead time s", "Trend bound s", "Bound check",
    ]
    title(ws, "Normal-process packet transport routes", len(headers))
    ws.cell(2, 1, "Effective inventory is formula-derived from design flow and the reduced-order 20 s anchor.")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.append([])
    ws.append(headers)
    for route_name, route in main.PROCESS_ROUTES.items():
        row = ws.max_row + 1
        ws.append([
            route_name, route.source, route.destination, route.design_carrier_kgh,
            route.design_dead_time_s, None, route.max_dead_time_s, 3600.0, None,
        ])
        ws.cell(row, 6, f"=D{row}*E{row}/3600")
        ws.cell(row, 9, f'=IF(E{row}<H{row},"PASS","FAIL")')
    style_table(ws, 4, 4, ws.max_row, len(headers))
    for row in range(5, ws.max_row + 1):
        for column in range(4, 9):
            ws.cell(row, column).number_format = "0.00"
    ws.conditional_formatting.add(
        f"I5:I{ws.max_row}", CellIsRule(operator="equal", formula=['"PASS"'], fill=PatternFill("solid", fgColor=GREEN))
    )
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:I{ws.max_row}"
    fit_columns(ws)


def add_summary(wb: Workbook, datasets: dict[str, dict]) -> None:
    ws = wb.active
    ws.title = "Summary"
    title(ws, "Urea OTS stream lag — evidence and selected model", 4)
    rows = [
        ("Finding", "Value", "Status", "Basis"),
        ("Normal independent anchors", f"=COUNTA('Normal Anchors'!A5:A{4 + len(datasets['Normal']['anchors'])})", "INFO", "17 hourly points"),
        ("Startup independent anchors", f"=COUNTA('Startup Anchors'!A5:A{4 + len(datasets['Startup']['anchors'])})", "INFO", "7 hourly points"),
        ("Subhour timing", "Not identifiable", "LIMIT", "30-second rows are synthetic linear interpolation"),
        ("Observed response bin", "0 h", "SUPPORTED", "Feed and several Unit 322 gradients move in the same hourly bin"),
        ("Defensible dead-time statement", "<3600 s", "SUPPORTED", "One-hour independent sampling interval"),
        ("Selected line anchor", "20 s at design liquid flow", "ASSUMPTION", "Existing reduced-order liquid-slug anchor"),
        ("Live route law", "theta = 3600 M_line / live mass flow", "MODEL", "Plug-flow effective inventory"),
        ("Receiver time constant", "tau approximately M / mass throughput", "MODEL", "Existing well-mixed vessel balances"),
        ("Property synchronization", "Whole packet", "REQUIRED", "Flow, temperature, composition, and Cp share one FIFO"),
    ]
    for row in rows:
        ws.append(row)
    style_table(ws, 2, 2, ws.max_row, 4)
    for row in range(3, ws.max_row + 1):
        status = ws.cell(row, 3).value
        ws.cell(row, 3).fill = PatternFill(
            "solid", fgColor=GREEN if status in {"SUPPORTED", "REQUIRED"} else RED if status == "LIMIT" else GRAY
        )
    ws.freeze_panes = "A3"
    fit_columns(ws, maximum=58)
    ws.column_dimensions["C"].width = 16


def build(trend_dir: Path, output: Path) -> Path:
    datasets: dict[str, dict] = {}
    for label, filename in TREND_FILES.items():
        notice, headers, anchors = hourly_anchors(trend_dir / filename)
        datasets[label] = {"notice": notice, "headers": headers, "anchors": anchors}

    wb = Workbook()
    for label, data in datasets.items():
        add_anchor_sheet(wb, label, data["headers"], data["anchors"])
        add_gradient_sheet(wb, label, data["headers"], len(data["anchors"]))
    add_lag_estimates(wb, datasets)
    add_route_parameters(wb)
    add_summary(wb, datasets)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def main_cli() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--trend-dir", type=Path, default=DEFAULT_TREND_DIR)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    result = build(args.trend_dir, args.output)
    print(result)


if __name__ == "__main__":
    main_cli()
