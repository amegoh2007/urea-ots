"""DCS System Identification — empirical θ (dead time) and τ (time constant) extraction.

Reads the two DCS trend Excel files, performs cross-correlation and step-response
analysis between manipulated variables (MV) and process variables (PV), then compares
the empirical dynamics against the simulator's current lag constants in main.py.

No pandas dependency — uses openpyxl + numpy + scipy directly.
"""

import math
import json
import os
import sys
from datetime import datetime, timedelta

import numpy as np
from scipy import signal as sig
from scipy.optimize import curve_fit
import openpyxl

# ============================================================================================
# 1. DATA LOADING
# ============================================================================================
DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "References", "DCS Trends")

def load_sheet1():
    """Urea_NormalOp_29-06-2025_Trends.xlsx — 30s interpolated, ~16h, Section 322 HP loop."""
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, "Urea_NormalOp_29-06-2025_Trends.xlsx"),
                                read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(min_row=3))   # row 1 = note, row 2 = blank, row 3 = headers
    headers = [c.value for c in rows[0]]
    data = {}
    for h in headers:
        data[h] = []
    for row in rows[1:]:
        vals = [c.value for c in row]
        if vals[0] is None:
            continue
        for h, v in zip(headers, vals):
            data[h].append(v)
    wb.close()
    # Convert timestamps to seconds from start
    t0 = data["Timestamp"][0]
    t_sec = []
    for ts in data["Timestamp"]:
        if isinstance(ts, datetime):
            t_sec.append((ts - t0).total_seconds())
        else:
            t_sec.append(0.0)
    data["t_sec"] = np.array(t_sec, dtype=float)
    for k in headers:
        if k != "Timestamp":
            arr = []
            for v in data[k]:
                try:
                    arr.append(float(v) if v is not None else np.nan)
                except (ValueError, TypeError):
                    arr.append(np.nan)
            data[k] = np.array(arr, dtype=float)
    return data, headers


def load_sheet2():
    """Book2.xlsx — ~2 months of DCS data at ~17min intervals, 145 tags."""
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, "Book2.xlsx"),
                                read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(min_row=1))
    headers = [c.value for c in rows[0]]
    data = {}
    for h in headers:
        if h is not None:
            data[h] = []
    ts_list = []
    for row in rows[1:]:
        vals = [c.value for c in row]
        ts_raw = vals[0]
        if ts_raw is None:
            continue
        ts_list.append(ts_raw)
        for h, v in zip(headers, vals):
            if h is not None and h in data:
                data[h].append(v)
    # Convert timestamps to seconds
    if ts_list:
        # timestamps are strings "YYYY-MM-DD HH:MM:SS"
        t0_str = ts_list[0]
        if isinstance(t0_str, str):
            t0 = datetime.strptime(t0_str, "%Y-%m-%d %H:%M:%S")
            t_sec = []
            for ts in ts_list:
                try:
                    dt_obj = datetime.strptime(str(ts), "%Y-%m-%d %H:%M:%S")
                    t_sec.append((dt_obj - t0).total_seconds())
                except:
                    t_sec.append(np.nan)
        elif isinstance(t0_str, datetime):
            t0 = t0_str
            t_sec = [(ts - t0).total_seconds() if isinstance(ts, datetime) else np.nan
                     for ts in ts_list]
        else:
            t_sec = list(range(len(ts_list)))
        data["t_sec"] = np.array(t_sec, dtype=float)
    # Convert numeric columns
    clean_headers = [h for h in headers if h is not None]
    for k in clean_headers:
        if k in data and k != headers[0]:
            arr = []
            for v in data[k]:
                try:
                    arr.append(float(v) if v is not None else np.nan)
                except (ValueError, TypeError):
                    arr.append(np.nan)
            data[k] = np.array(arr, dtype=float)
    wb.close()
    return data, clean_headers


# ============================================================================================
# 2. SYSTEM IDENTIFICATION FUNCTIONS
# ============================================================================================

def detrend_and_normalize(x):
    """Remove NaN, detrend, normalize to zero-mean unit-variance."""
    mask = ~np.isnan(x)
    y = x.copy()
    y[~mask] = np.nanmean(x)  # fill NaN with mean
    y = y - np.mean(y)
    s = np.std(y)
    if s > 1e-12:
        y = y / s
    return y


def cross_corr_lag(mv, pv, dt_sample, max_lag_s=3600):
    """Estimate dead time by cross-correlation peak between MV and PV signals.
    
    Returns:
        theta_s: dead time in seconds (positive = PV lags MV)
        corr_peak: peak cross-correlation coefficient
        lags_s: lag axis (seconds)
        corr: full cross-correlation array
    """
    mv_n = detrend_and_normalize(mv)
    pv_n = detrend_and_normalize(pv)
    n = len(mv_n)
    max_lag_samples = min(int(max_lag_s / dt_sample), n - 1)
    
    corr = np.correlate(pv_n, mv_n, mode='full')
    corr = corr / n  # normalize
    mid = n - 1
    # Only look at positive lags (PV lags MV)
    lags = np.arange(-mid, mid + 1) * dt_sample
    
    # Search positive lags only (up to max_lag)
    pos_start = mid
    pos_end = min(mid + max_lag_samples, len(corr) - 1)
    pos_corr = corr[pos_start:pos_end + 1]
    pos_lags = lags[pos_start:pos_end + 1]
    
    if len(pos_corr) == 0:
        return 0.0, 0.0, lags, corr
    
    peak_idx = np.argmax(pos_corr)
    theta_s = pos_lags[peak_idx]
    corr_peak = pos_corr[peak_idx]
    
    return theta_s, corr_peak, pos_lags, pos_corr


def fit_foptd_step(t, pv, theta_guess, tau_guess):
    """Fit a First-Order-Plus-Dead-Time model to a step-response segment.
    
    Model: y(t) = K * (1 - exp(-(t - theta) / tau)) for t >= theta, else 0.
    
    Returns: K, tau, theta (fitted)
    """
    pv_0 = pv[0]
    pv_ss = np.mean(pv[-max(1, len(pv)//10):])  # last 10% as steady state
    K_guess = pv_ss - pv_0
    
    if abs(K_guess) < 1e-12:
        return 0.0, tau_guess, theta_guess
    
    def foptd_model(t_arr, K, tau, theta):
        y = np.zeros_like(t_arr, dtype=float)
        for i, ti in enumerate(t_arr):
            if ti >= theta and tau > 0:
                y[i] = pv_0 + K * (1.0 - math.exp(-(ti - theta) / tau))
            else:
                y[i] = pv_0
        return y
    
    try:
        popt, pcov = curve_fit(
            foptd_model, t - t[0], pv,
            p0=[K_guess, max(tau_guess, 1.0), max(theta_guess, 0.0)],
            bounds=([K_guess * 0.1 if K_guess > 0 else K_guess * 5,
                     0.1, 0.0],
                    [K_guess * 5 if K_guess > 0 else K_guess * 0.1,
                     max_lag_s := 7200.0, 3600.0]),
            maxfev=5000
        )
        return popt  # K, tau, theta
    except Exception:
        return K_guess, tau_guess, theta_guess


def estimate_time_constant_63pct(t, pv, theta_s=0.0):
    """Estimate τ from the 63.2% rise time method (Smith's method).
    
    After accounting for dead time, finds when the signal reaches 63.2% of
    its final value change.
    """
    pv_0 = pv[0]
    pv_ss = np.mean(pv[-max(1, len(pv)//10):])
    delta = pv_ss - pv_0
    if abs(delta) < 1e-12:
        return 0.0
    
    target = pv_0 + 0.632 * delta
    
    # Find first crossing after dead time
    t_start = t[0] + theta_s
    for i in range(len(t)):
        if t[i] < t_start:
            continue
        if delta > 0 and pv[i] >= target:
            tau = t[i] - t_start
            return tau
        elif delta < 0 and pv[i] <= target:
            tau = t[i] - t_start
            return tau
    return t[-1] - t_start  # never reached


def compute_process_gain(mv, pv):
    """Static process gain K_p = Δ(PV_ss) / Δ(MV_ss)."""
    mv_ss_start = np.mean(mv[:max(1, len(mv)//10)])
    mv_ss_end = np.mean(mv[-max(1, len(mv)//10):])
    pv_ss_start = np.mean(pv[:max(1, len(pv)//10)])
    pv_ss_end = np.mean(pv[-max(1, len(pv)//10):])
    
    dmv = mv_ss_end - mv_ss_start
    dpv = pv_ss_end - pv_ss_start
    
    if abs(dmv) < 1e-12:
        return 0.0
    return dpv / dmv


# ============================================================================================
# 3. MV-PV PAIR DEFINITIONS
# ============================================================================================
# Maps of (MV_tag, PV_tag, description, expected_lag_mechanism, simulator_tau_constant_name)
# organized by plant section.

PAIRS_SHEET1 = [
    # Section 322 — HP Synthesis Loop
    ("HIC-322605", "TT-322014",
     "HV-322605 spindle → reactor overflow T",
     "thermal inertia (reactor liquid holdup + metal mass)",
     "REACT_THERM_TAU_MIN * 60 = 480"),
    ("HIC-322605", "TT-322004",
     "HV-322605 spindle → stripper bottom T",
     "thermal inertia (stripper falling-film + sump + shell metal)",
     "STRIP_T_TAU_S = 180"),
    ("HIC-322605", "TT-322013",
     "HV-322605 spindle → stripper top T",
     "thermal inertia (stripper reflux + tube bundle)",
     "STRIP_T_TAU_S = 180"),
    ("HIC-322604", "TT-322011",
     "HV-322604 off-gas valve → off-gas T",
     "thermal inertia (offgas line + thermowell metal)",
     "OFFGAS_T_TAU_S = 120"),
    ("HIC-322604", "TT-322002",
     "HV-322604 off-gas valve → scrubber overflow T",
     "thermal inertia (scrubber liquid pool + shell metal)",
     "SCRUB_T_TAU_S = 180"),
    ("SIC-321951", "TT-322014",
     "NH₃ pump speed → reactor overflow T",
     "transport delay (NH₃ pipe + reactor residence) + thermal",
     "FEED_TD_S=345 + REACT_TAU_REC_MIN*60=300"),
    ("HV-322602", "PT-329206",
     "HV-322602 steam valve → LP header P",
     "header capacitance (322D001A/B)",
     "C_LP = 25.0 → τ ≈ C_LP/flow"),
    ("LV-322501", "TT-322004",
     "LV-322501 drain → stripper bottom T",
     "thermal inertia (flash + sump + pipe)",
     "STRIP_T_TAU_S = 180"),
    ("UREA-LOAD", "TT-322005",
     "Load % → reactor liquid T (TT-322005)",
     "synthesis loop recycle lag",
     "REACT_TAU_REC_MIN * 60 = 300"),
    ("UREA-LOAD", "AY-322701",
     "Load % → N/C analyzer reading",
     "analyzer sampling + measurement lag",
     "AT_322701_TAU_S = 40"),
    ("HIC-322605", "TT-322010",
     "HV-322605 spindle → HPCC product T",
     "thermal inertia (HPCC tube bundle + liquid holdup)",
     "HPCC_T_TAU_S = 240"),
    ("PIC-329204", "TT-322012",
     "HP steam chest P → ejector discharge T",
     "thermal inertia (ejector + suction carbamate)",
     "EJ_T_TAU_S = 120"),
    ("FYM-322403", "TT-322005",
     "CO₂ feed flow → reactor liquid T",
     "transport (CO₂ pipe) + reactor thermal inertia",
     "FEED_TD_S=345 + thermal"),
    ("TIC-329005", "TDY-329125",
     "CCW temp controller → CCW return ΔT",
     "tempered water shell return lag",
     "CCW_T_TAU_S = 25"),
    ("PIC-322203", "PV-322203",
     "CO₂ line pressure → PV-322203 valve stroke",
     "controller response (direct-acting PIC)",
     "controller only (no process lag)"),
]

PAIRS_SHEET2 = [
    # Section 323 — MP Decomposition
    ("LIC-322501", "TT-323001",
     "LV-322501 drain → column feed T (TT-323001)",
     "flash lag (LV → column entry)",
     "STRIP_T_TAU_S = 180"),
    ("LIC-322501", "TT-323002",
     "LV-322501 drain → column sump T (TT-323002)",
     "holdup lag (column liquid residence)",
     "R323_C003_M_TAU_S = 120"),
    ("PIC-329202", "TT-323004",
     "MP steam chest P → C003 column T",
     "thermal inertia (reboiler + column holdup)",
     "R323_C003_M_TAU_S = 120"),
    ("FIC-323401", "TT-323005",
     "LP steam to E010 → pre-evap T",
     "thermal inertia (323E010 shell + holdup)",
     "R323_F010_M_TAU_S = 240"),
    ("PIC-323203", "PT-323201",
     "Flash drum PIC → column overhead P",
     "gas holdup (flash vaporization + piping)",
     "R323_F004_P_TAU_S = 90"),
    # Section 328 — LP Decomposition
    ("TIC-328002", "TT-328004",
     "328 temp controller → LP decomposer T",
     "thermal inertia (decomposer compartments)",
     "a328_c001_T: implicit holdup ODE"),
    ("PIC-328202", "FIC-328402",
     "328 pressure → carbamate return flow",
     "hydraulic lag (level + pump response)",
     "None — implicit level dynamics"),
    # Section 329 — Steam
    ("FIC-329401", "PT-329207",
     "BL steam makeup → LP header P",
     "header capacitance",
     "C_LP = 25.0"),
    ("PV-329204", "PT-329201",
     "PV-329204 stroke → HP steam drum P",
     "header capacitance (329D005)",
     "C_MP = 25.0"),
    ("TIC-329005", "TT-329004",
     "CCW temp controller → cooling water T",
     "shell-side thermal mass",
     "CCW_T_TAU_S = 25"),
    # Section 322 (additional from Book2)
    ("HVGT-322605", "TT-322014",
     "HV-322605 travel → reactor overflow T",
     "thermal inertia (reactor liquid holdup)",
     "REACT_THERM_TAU_MIN * 60 = 480"),
    ("LIC-322501", "TT-322004",
     "LV-322501 → stripper bottom T",
     "thermal inertia (flash + sump)",
     "STRIP_T_TAU_S = 180"),
    ("UREA-LOAD", "AY-322701",
     "Load → N/C analyzer",
     "analyzer sampling + measurement lag",
     "AT_322701_TAU_S = 40"),
    ("SIC-321951", "TT-322002",
     "NH₃ pump speed → scrubber overflow T",
     "transport + thermal (loop recycle + scrubber)",
     "FEED_TD_S=345 + SCRUB_T_TAU_S=180"),
    ("SIC-321950", "PT-321202",
     "NH₃ pump A speed → discharge P",
     "pump acceleration + piping",
     "instantaneous (PD pump, no lag expected)"),
    ("FIC-320402", "TT-322017",
     "CO₂ feed control → 322 line T",
     "transport delay (feed pipe + compressor)",
     "FEED_TD_S = 345"),
    ("PIC-322201", "TT-322015",
     "Absorber pressure → absorber off-gas T",
     "gas holdup lag (322C001 volume)",
     "R323_C003_P_TAU_S = 1"),
    ("TIC-322021", "TT-322002",
     "Temp controller → scrubber overflow T",
     "thermal (scrubber inventory + CCW shell)",
     "SCRUB_T_TAU_S = 180"),
]


# ============================================================================================
# 4. ANALYSIS ENGINE
# ============================================================================================

def analyze_pair(data, mv_tag, pv_tag, description, expected, sim_tau_str, dt_sample):
    """Run cross-correlation and time-constant estimation for one MV→PV pair."""
    if mv_tag not in data or pv_tag not in data:
        return None
    mv = data[mv_tag]
    pv = data[pv_tag]
    t = data["t_sec"]
    
    if not isinstance(mv, np.ndarray) or not isinstance(pv, np.ndarray):
        return None
    if len(mv) < 20 or len(pv) < 20:
        return None
    
    # Check for sufficient variation
    mv_std = np.nanstd(mv)
    pv_std = np.nanstd(pv)
    if mv_std < 1e-6 or pv_std < 1e-6:
        return {
            "mv_tag": mv_tag, "pv_tag": pv_tag,
            "description": description,
            "status": "INSUFFICIENT_VARIATION",
            "mv_std": float(mv_std), "pv_std": float(pv_std),
            "sim_tau_str": sim_tau_str,
        }
    
    # Cross-correlation for dead time
    theta_s, corr_peak, lags, corr = cross_corr_lag(mv, pv, dt_sample, max_lag_s=7200)
    
    # Time constant via autocorrelation decay of PV
    pv_n = detrend_and_normalize(pv)
    acf = np.correlate(pv_n, pv_n, mode='full')
    acf = acf[len(pv_n)-1:]  # keep positive lags only
    acf = acf / acf[0]
    # Find where ACF drops below 1/e ≈ 0.368
    tau_acf_s = 0.0
    for i in range(len(acf)):
        if acf[i] < 1.0/math.e:
            tau_acf_s = i * dt_sample
            break
    else:
        tau_acf_s = len(acf) * dt_sample  # never dropped below 1/e
    
    # Process gain (static)
    K_p = compute_process_gain(mv, pv)
    
    return {
        "mv_tag": mv_tag,
        "pv_tag": pv_tag,
        "description": description,
        "expected_mechanism": expected,
        "sim_tau_str": sim_tau_str,
        "theta_s": float(theta_s),
        "tau_acf_s": float(tau_acf_s),
        "K_p": float(K_p),
        "xcorr_peak": float(corr_peak),
        "mv_mean": float(np.nanmean(mv)),
        "mv_std": float(mv_std),
        "pv_mean": float(np.nanmean(pv)),
        "pv_std": float(pv_std),
        "dt_sample": float(dt_sample),
        "n_points": int(len(mv)),
    }


# ============================================================================================
# 5. MAIN ANALYSIS
# ============================================================================================

def run_analysis():
    print("=" * 80)
    print("  DCS SYSTEM IDENTIFICATION — Empirical θ and τ Extraction")
    print("=" * 80)
    
    results = {"sheet1": [], "sheet2": []}
    
    # --- Sheet 1: Normal Operation (30s intervals) ---
    print("\n▶ Loading Sheet 1: Urea_NormalOp_29-06-2025_Trends.xlsx ...")
    data1, hdrs1 = load_sheet1()
    dt1 = 30.0  # 30s interpolated
    n1 = len(data1["t_sec"])
    t_span1 = data1["t_sec"][-1] - data1["t_sec"][0]
    print(f"  Tags: {len(hdrs1)}, Samples: {n1}, Span: {t_span1/3600:.1f} h, Δt: {dt1} s")
    print(f"  Available tags: {[h for h in hdrs1 if h != 'Timestamp']}")
    
    print("\n  Analyzing MV→PV pairs (Sheet 1) ...")
    for mv_tag, pv_tag, desc, expected, sim_tau in PAIRS_SHEET1:
        r = analyze_pair(data1, mv_tag, pv_tag, desc, expected, sim_tau, dt1)
        if r:
            results["sheet1"].append(r)
            status = r.get("status", "OK")
            if status == "OK":
                print(f"    ✓ {mv_tag:16s} → {pv_tag:12s}  θ={r['theta_s']:7.1f}s  τ_acf={r['tau_acf_s']:7.1f}s  K={r['K_p']:+8.4f}  r={r['xcorr_peak']:.3f}")
            else:
                print(f"    ✗ {mv_tag:16s} → {pv_tag:12s}  {status}")
    
    # --- Sheet 2: Two-month data (~17min intervals) ---
    print("\n▶ Loading Sheet 2: Book2.xlsx ...")
    data2, hdrs2 = load_sheet2()
    # Compute actual dt from timestamps
    t2 = data2["t_sec"]
    valid_t2 = t2[~np.isnan(t2)]
    if len(valid_t2) > 1:
        dt2 = np.median(np.diff(valid_t2))
    else:
        dt2 = 1008.0  # fallback (~17 min)
    n2 = len(valid_t2)
    t_span2 = valid_t2[-1] - valid_t2[0]
    print(f"  Tags: {len(hdrs2)}, Samples: {n2}, Span: {t_span2/86400:.1f} days, Δt: {dt2:.0f} s ({dt2/60:.1f} min)")
    
    print("\n  Analyzing MV→PV pairs (Sheet 2) ...")
    for mv_tag, pv_tag, desc, expected, sim_tau in PAIRS_SHEET2:
        r = analyze_pair(data2, mv_tag, pv_tag, desc, expected, sim_tau, dt2)
        if r:
            results["sheet2"].append(r)
            status = r.get("status", "OK")
            if status == "OK":
                print(f"    ✓ {mv_tag:16s} → {pv_tag:12s}  θ={r['theta_s']:7.1f}s  τ_acf={r['tau_acf_s']:7.1f}s  K={r['K_p']:+8.4f}  r={r['xcorr_peak']:.3f}")
            else:
                print(f"    ✗ {mv_tag:16s} → {pv_tag:12s}  {status}")
    
    return results


def generate_comparison_report(results):
    """Generate the comparison report between empirical and simulator dynamics."""
    
    # ============================================================================================
    # Current simulator tau constants (extracted from main.py)
    # ============================================================================================
    SIM_TAUS = {
        # Section 322 display lags
        "EJ_T_TAU_S": 120.0,
        "STRIP_T_TAU_S": 180.0,
        "HPCC_T_TAU_S": 240.0,
        "HPCC_P_TAU_S": 30.0,
        "SCRUB_T_TAU_S": 180.0,
        "OFFGAS_T_TAU_S": 120.0,
        "CCW_T_TAU_S": 25.0,
        "AT_322701_TAU_S": 40.0,
        "SCRUB_LVL_TAU_S": 120.0,
        # Feed transport delay
        "FEED_TD_S": 345.0,
        # Section 323 dynamics
        "R323_C003_M_TAU_S": 120.0,
        "R323_C003_P_TAU_S": 1.0,
        "R323_F004_M_TAU_S": 180.0,
        "R323_F004_P_TAU_S": 90.0,
        "R323_F010_M_TAU_S": 240.0,
        # Reactor
        "REACT_THERM_TAU_MIN": 8.0,  # min → 480 s
        "REACT_TAU_REC_MIN": 5.0,    # min → 300 s
        "REACT_FWD_TAU_MIN": 8.0,    # min → 480 s
        # Section 324
        "R324_F001_M_TAU_S": 180.0,
        "R324_E001_COND_TAU_S": 90.0,
        "R324_F003_M_TAU_S": 180.0,
        # Section 329
        "C_MP": 25.0,  # (kg/s)/bar (header capacitance)
        "C_LP": 25.0,
        "C_9": 53.2,
        # 322E001 bottom lag
        "strip_bot_kgh_delay": 60.0,
        # Scrubber holdup
        "SCRUB_TAU_HOLDUP_MIN": 4.0,  # min → 240 s
        # HPCC fill
        "HPCC_TAU_FILL_MIN": 6.0,    # min → 360 s
    }
    
    lines = []
    lines.append("=" * 100)
    lines.append("  DCS SYSTEM IDENTIFICATION REPORT — Simulator Calibration Audit")
    lines.append("=" * 100)
    lines.append("")
    
    # Collect all results
    all_results = results.get("sheet1", []) + results.get("sheet2", [])
    
    # Organize by physical mechanism
    categories = {
        "HP Synthesis Loop Temperatures (Section 322)": [],
        "Reactor Thermal Dynamics": [],
        "MP Decomposition (Section 323)": [],
        "LP Decomposition (Section 328)": [],
        "Steam Network (Section 329)": [],
        "Feed Transport Delays": [],
        "Analyzer & Instrument Dynamics": [],
        "Controller Response": [],
    }
    
    for r in all_results:
        if r.get("status") == "INSUFFICIENT_VARIATION":
            continue
        desc = r["description"].lower()
        if "reactor" in desc and "overflow" in desc:
            categories["Reactor Thermal Dynamics"].append(r)
        elif "analyzer" in desc or "n/c" in desc:
            categories["Analyzer & Instrument Dynamics"].append(r)
        elif "322" in r["pv_tag"] or "stripper" in desc or "scrubber" in desc or "ejector" in desc or "hpcc" in desc:
            categories["HP Synthesis Loop Temperatures (Section 322)"].append(r)
        elif "323" in r["pv_tag"] or "column" in desc or "pre-evap" in desc:
            categories["MP Decomposition (Section 323)"].append(r)
        elif "328" in r["pv_tag"] or "decomposer" in desc:
            categories["LP Decomposition (Section 328)"].append(r)
        elif "329" in r["pv_tag"] or "steam" in desc or "header" in desc:
            categories["Steam Network (Section 329)"].append(r)
        elif "feed" in desc or "co₂" in desc.lower() or "co2" in desc.lower():
            categories["Feed Transport Delays"].append(r)
        elif "controller" in desc:
            categories["Controller Response"].append(r)
        elif "reactor" in desc:
            categories["Reactor Thermal Dynamics"].append(r)
        else:
            # Fallback: assign by section number in tag
            if "322" in r["pv_tag"]:
                categories["HP Synthesis Loop Temperatures (Section 322)"].append(r)
            elif "323" in r["pv_tag"]:
                categories["MP Decomposition (Section 323)"].append(r)
            elif "328" in r["pv_tag"]:
                categories["LP Decomposition (Section 328)"].append(r)
            elif "329" in r["pv_tag"]:
                categories["Steam Network (Section 329)"].append(r)
            else:
                categories["Controller Response"].append(r)
    
    discrepancies = []
    
    for cat_name, cat_results in categories.items():
        if not cat_results:
            continue
        lines.append(f"\n{'─' * 100}")
        lines.append(f"  {cat_name}")
        lines.append(f"{'─' * 100}")
        lines.append(f"{'MV Tag':<18s} {'PV Tag':<14s} {'θ_emp (s)':>10s} {'τ_emp (s)':>10s} {'K_p':>9s} "
                      f"{'r_xcorr':>8s} {'Sim τ/θ':>12s} {'Verdict':>12s}")
        lines.append("─" * 100)
        
        for r in cat_results:
            theta = r["theta_s"]
            tau = r["tau_acf_s"]
            Kp = r["K_p"]
            rx = r["xcorr_peak"]
            sim_str = r["sim_tau_str"]
            
            # Parse simulator tau for comparison
            # Extract numbers from sim_tau_str
            import re
            nums = re.findall(r'[\d.]+', sim_str)
            sim_tau_val = float(nums[0]) if nums else 0.0
            # Adjust if the constant is in minutes
            if "MIN" in sim_str.upper() and "*" in sim_str and "60" in sim_str:
                pass  # already converted in the string
            
            # Simple ratio-based verdict
            if sim_tau_val > 0:
                ratio = tau / sim_tau_val
                if 0.5 <= ratio <= 2.0:
                    verdict = "OK"
                elif ratio < 0.5:
                    verdict = "SIM_SLOW"
                else:
                    verdict = "SIM_FAST"
            else:
                verdict = "N/A"
            
            if verdict in ("SIM_SLOW", "SIM_FAST"):
                discrepancies.append({
                    "mv_tag": r["mv_tag"],
                    "pv_tag": r["pv_tag"],
                    "description": r["description"],
                    "theta_emp": theta,
                    "tau_emp": tau,
                    "sim_tau_str": sim_str,
                    "sim_tau_val": sim_tau_val,
                    "verdict": verdict,
                    "ratio": ratio if sim_tau_val > 0 else 0.0,
                    "mechanism": r.get("expected_mechanism", ""),
                })
            
            lines.append(f"{r['mv_tag']:<18s} {r['pv_tag']:<14s} {theta:10.1f} {tau:10.1f} {Kp:+9.4f} "
                          f"{rx:8.3f} {sim_str:>12s} {verdict:>12s}")
    
    lines.append(f"\n{'=' * 100}")
    lines.append(f"  DISCREPANCY SUMMARY — {len(discrepancies)} parameter(s) require update")
    lines.append(f"{'=' * 100}")
    
    if discrepancies:
        for d in discrepancies:
            lines.append(f"\n  ▸ {d['mv_tag']} → {d['pv_tag']}: {d['description']}")
            lines.append(f"    Empirical: θ = {d['theta_emp']:.1f} s,  τ = {d['tau_emp']:.1f} s")
            lines.append(f"    Simulator: {d['sim_tau_str']}")
            if d["sim_tau_val"] > 0:
                lines.append(f"    Ratio τ_emp/τ_sim = {d['ratio']:.2f}  →  {'Sim too slow' if d['verdict'] == 'SIM_SLOW' else 'Sim too fast'}")
            lines.append(f"    Mechanism: {d['mechanism']}")
    else:
        lines.append("\n  All analyzed dynamics are within 2× of simulator constants. No corrections needed.")
    
    report_text = "\n".join(lines)
    return report_text, discrepancies


# ============================================================================================
# 6. ENTRY POINT
# ============================================================================================

if __name__ == "__main__":
    results = run_analysis()
    report, discrepancies = generate_comparison_report(results)
    print(report)
    
    # Save JSON results
    out_dir = os.path.join(os.path.dirname(__file__), "reports")
    os.makedirs(out_dir, exist_ok=True)
    
    # Save raw results
    json_path = os.path.join(out_dir, "dcs_sysid_results.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, default=str)
    print(f"\n  Raw results saved to: {json_path}")
    
    # Save report
    report_path = os.path.join(out_dir, "dcs_sysid_report.md")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report)
    print(f"  Report saved to: {report_path}")
    
    # Save discrepancies
    disc_path = os.path.join(out_dir, "dcs_sysid_discrepancies.json")
    with open(disc_path, "w", encoding="utf-8") as f:
        json.dump(discrepancies, f, indent=2, default=str)
    print(f"  Discrepancies saved to: {disc_path}")
